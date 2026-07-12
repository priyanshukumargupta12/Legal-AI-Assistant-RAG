"""
app/dataset/dataset_repository.py
====================================
Repository layer for the Dataset Management Module.

PURPOSE:
    Abstracts all persistence operations for the dataset module:
    reading from the file system, writing CSV, writing Excel, writing JSON.

    The abstract class DatasetRepository defines the contract.
    FileSystemDatasetRepository implements it using pandas + openpyxl.

WHY REPOSITORY PATTERN:
    - Services depend only on the abstract interface.
    - If storage changes (e.g., switch from file-based to SQLite), only
      the concrete class changes — zero changes to DatasetService.
    - The abstract interface is mockable in unit tests.
    - Clean separation: service owns logic, repository owns I/O.

DESIGN:
    - DatasetRepository:          Abstract base class (ABC)
    - FileSystemDatasetRepository: Writes CSV, XLSX, JSON to metadata/
    - Professional XLSX formatting via openpyxl:
        • Frozen header row
        • Bold + colored header
        • Status column color coding
        • Auto-width columns

SOLID:
    Single Responsibility — only handles persistence of dataset metadata.
    Dependency Inversion   — DatasetService depends on the abstract class.
    Open/Closed            — add S3Repository without touching the service.
"""

from __future__ import annotations

import json
from abc import ABC, abstractmethod
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.core.constants import (
    DOCUMENTS_CSV_FILENAME,
    DOCUMENTS_XLSX_FILENAME,
    DOC_STATUS_VALID,
    DOC_STATUS_INVALID,
    DOC_STATUS_DUPLICATE,
)
from app.core.exceptions import DatasetScanError
from app.dataset.dataset_logger import dataset_log
from app.dataset.dataset_models import DatasetStatistics, DocumentRecord, DocumentStatus, ScanResult


# =============================================================================
# ABSTRACT REPOSITORY INTERFACE
# =============================================================================

class DatasetRepository(ABC):
    """
    Abstract contract for dataset metadata persistence.

    All concrete implementations must implement every abstract method.
    DatasetService depends only on this interface.
    """

    @abstractmethod
    def save_documents_csv(
        self,
        documents: List[DocumentRecord],
        output_path: Path,
    ) -> Path:
        """
        Persist document records to a CSV file.

        Args:
            documents:   List of DocumentRecord objects to serialize.
            output_path: File path where the CSV should be written.

        Returns:
            Resolved absolute path of the written CSV file.

        Raises:
            DatasetScanError: If the file cannot be written.
        """
        ...

    @abstractmethod
    def save_documents_xlsx(
        self,
        documents: List[DocumentRecord],
        statistics: DatasetStatistics,
        output_path: Path,
    ) -> Path:
        """
        Persist document records to a professionally formatted Excel file.

        Args:
            documents:   List of DocumentRecord objects.
            statistics:  Aggregate stats (written to a second 'Summary' sheet).
            output_path: File path where the XLSX should be written.

        Returns:
            Resolved absolute path of the written XLSX file.

        Raises:
            DatasetScanError: If the file cannot be written.
        """
        ...

    @abstractmethod
    def save_json_summary(
        self,
        statistics: DatasetStatistics,
        output_path: Path,
    ) -> Path:
        """
        Persist dataset statistics to a JSON summary file.

        Args:
            statistics:  DatasetStatistics domain model.
            output_path: File path where the JSON should be written.

        Returns:
            Resolved absolute path of the written JSON file.

        Raises:
            DatasetScanError: If the file cannot be written.
        """
        ...

    @abstractmethod
    def load_documents_csv(self, csv_path: Path) -> List[dict]:
        """
        Load previously saved document records from a CSV file.

        Args:
            csv_path: Path to the CSV file.

        Returns:
            List of raw row dicts (one per CSV row).

        Raises:
            FileNotFoundError:  If csv_path does not exist.
            DatasetScanError:   If the CSV cannot be parsed.
        """
        ...


# =============================================================================
# CONCRETE IMPLEMENTATION — FILE SYSTEM
# =============================================================================

class FileSystemDatasetRepository(DatasetRepository):
    """
    File-system implementation of DatasetRepository.

    Writes metadata to metadata/ directory using:
        - pandas: DataFrame construction and CSV/XLSX writing
        - openpyxl: Professional XLSX formatting (colors, borders, widths)
        - json: Standard library JSON serialization

    The metadata/ directory is created if it does not exist.
    """

    # ── CSV column order (matches specification) ──────────────────────────────
    _CSV_COLUMNS: List[str] = [
        "Document_ID",
        "Category",
        "File_Name",
        "Title",
        "File_Path",
        "Relative_Path",
        "Pages",
        "File_Size_Bytes",
        "File_Size_MB",
        "Checksum_SHA256",
        "Status",
        "Created_At",
        "Modified_At",
        "Error_Message",
        "Is_Duplicate_Of",
    ]

    # ── XLSX formatting constants ─────────────────────────────────────────────
    _HEADER_FILL = PatternFill(
        start_color="1E3A5F",  # Deep navy blue
        end_color="1E3A5F",
        fill_type="solid",
    )
    _HEADER_FONT = Font(
        name="Calibri",
        bold=True,
        color="FFFFFF",  # White text
        size=11,
    )
    _THIN_BORDER_SIDE = Side(style="thin", color="CCCCCC")
    _CELL_BORDER = Border(
        left=_THIN_BORDER_SIDE,
        right=_THIN_BORDER_SIDE,
        top=_THIN_BORDER_SIDE,
        bottom=_THIN_BORDER_SIDE,
    )

    # Status → fill color mapping for the Status column
    _STATUS_FILLS: dict[str, PatternFill] = {
        "valid":     PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "invalid":   PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "corrupted": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "empty":     PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "duplicate": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    }

    # ── Initializer ───────────────────────────────────────────────────────────

    def __init__(self, metadata_dir: Path) -> None:
        """
        Initialize the repository with the target metadata directory.

        Args:
            metadata_dir: Directory where output files will be written.
                          Created automatically if it does not exist.
        """
        self._metadata_dir = metadata_dir.resolve()
        self._metadata_dir.mkdir(parents=True, exist_ok=True)
        dataset_log.debug(
            "FileSystemDatasetRepository initialized | metadata_dir={path}",
            path=str(self._metadata_dir),
        )

    # ── Private helpers ───────────────────────────────────────────────────────

    def _build_dataframe(self, documents: List[DocumentRecord]) -> pd.DataFrame:
        """
        Convert a list of DocumentRecord objects to a pandas DataFrame.

        The DataFrame column order matches _CSV_COLUMNS exactly.

        Args:
            documents: List of DocumentRecord domain objects.

        Returns:
            pandas DataFrame with one row per document.
        """
        rows = [
            {
                "Document_ID":      doc.document_id,
                "Category":         doc.category,
                "File_Name":        doc.file_name,
                "Title":            doc.title,
                "File_Path":        doc.file_path,
                "Relative_Path":    doc.relative_path,
                "Pages":            doc.page_count,
                "File_Size_Bytes":  doc.file_size_bytes,
                "File_Size_MB":     doc.file_size_mb,
                "Checksum_SHA256":  doc.checksum_sha256,
                "Status":           doc.status.value,
                "Created_At":       doc.created_at.isoformat(),
                "Modified_At":      doc.modified_at.isoformat(),
                "Error_Message":    doc.error_message or "",
                "Is_Duplicate_Of":  doc.is_duplicate_of or "",
            }
            for doc in documents
        ]

        if not rows:
            # Return empty DataFrame with correct columns
            return pd.DataFrame(columns=self._CSV_COLUMNS)

        return pd.DataFrame(rows, columns=self._CSV_COLUMNS)

    def _apply_xlsx_header_formatting(self, worksheet) -> None:
        """
        Apply professional header formatting to the first row of a worksheet.

        Applies: navy blue fill, white bold Calibri font, centred alignment,
        and thin borders.

        Args:
            worksheet: openpyxl Worksheet object.
        """
        for cell in worksheet[1]:
            cell.fill = self._HEADER_FILL
            cell.font = self._HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = self._CELL_BORDER

    def _apply_xlsx_status_colors(self, worksheet, status_col_idx: int) -> None:
        """
        Color-code cells in the Status column based on their value.

        Green  → valid
        Red    → invalid | corrupted
        Yellow → empty | duplicate

        Args:
            worksheet:      openpyxl Worksheet object.
            status_col_idx: 1-based column index of the Status column.
        """
        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=2, max_row=worksheet.max_row),
            start=2,
        ):
            status_cell = row[status_col_idx - 1]  # Convert to 0-based list index
            status_value = str(status_cell.value or "").lower()
            fill = self._STATUS_FILLS.get(status_value)
            if fill:
                status_cell.fill = fill

    def _apply_xlsx_column_widths(self, worksheet) -> None:
        """
        Set column widths based on the maximum content length in each column.

        Caps width at 50 characters to prevent extremely wide columns for
        long file paths.

        Args:
            worksheet: openpyxl Worksheet object.
        """
        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            # Add padding; cap at 50
            adjusted_width = min(max_length + 4, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    def _apply_xlsx_row_formatting(self, worksheet, status_col_idx: int) -> None:
        """
        Apply borders and alignment to all data rows (row 2 onwards).

        Args:
            worksheet:      openpyxl Worksheet.
            status_col_idx: 1-based Status column index.
        """
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = self._CELL_BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _add_statistics_sheet(
        self,
        workbook,
        statistics: DatasetStatistics,
    ) -> None:
        """
        Add a 'Summary' sheet to the workbook with aggregate statistics.

        Args:
            workbook:   openpyxl Workbook object.
            statistics: DatasetStatistics domain model.
        """
        ws = workbook.create_sheet(title="Summary")

        summary_data = [
            ("Metric", "Value"),
            ("Scan Time", statistics.scanned_at.strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Dataset Root", statistics.dataset_root),
            ("", ""),
            ("DOCUMENT COUNTS", ""),
            ("Total Documents", statistics.total_documents),
            ("Valid Documents", statistics.valid_documents),
            ("Invalid Documents", statistics.invalid_documents),
            ("Empty Documents", statistics.empty_documents),
            ("Duplicate Documents", statistics.duplicate_documents),
            ("Valid %", f"{statistics.valid_percentage:.1f}%"),
            ("", ""),
            ("SIZE STATISTICS", ""),
            ("Total Size (MB)", f"{statistics.total_size_mb:.2f} MB"),
            ("Total Size (GB)", f"{statistics.total_size_gb:.4f} GB"),
            ("Average File Size (MB)", f"{statistics.avg_file_size_mb:.3f} MB"),
            ("", ""),
            ("PAGE STATISTICS", ""),
            ("Average Pages", f"{statistics.avg_pages:.1f}"),
            ("Largest PDF", statistics.largest_pdf_name or "N/A"),
            ("Smallest PDF", statistics.smallest_pdf_name or "N/A"),
            ("", ""),
            ("CATEGORY BREAKDOWN", ""),
        ]

        for cat, count in statistics.category_counts.items():
            summary_data.append((f"  {cat}", count))

        # Write rows
        for row_idx, (label, value) in enumerate(summary_data, start=1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

            # Format header row and section headers
            if row_idx == 1 or str(label).isupper():
                for col in (1, 2):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True, color="1E3A5F")
                    cell.fill = PatternFill(
                        start_color="E8F0F7", end_color="E8F0F7", fill_type="solid"
                    )

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30

    # ── Public interface ──────────────────────────────────────────────────────

    def save_documents_csv(
        self,
        documents: List[DocumentRecord],
        output_path: Optional[Path] = None,
    ) -> Path:
        """
        Write document records to a CSV file.

        Args:
            documents:   List of DocumentRecord objects to export.
            output_path: Target CSV path. Defaults to metadata/documents.csv.

        Returns:
            Resolved absolute path of the written CSV file.

        Raises:
            DatasetScanError: If pandas CSV write fails.
        """
        target = (output_path or self._metadata_dir / DOCUMENTS_CSV_FILENAME).resolve()

        dataset_log.info(
            "Writing CSV | path={path} | records={count}",
            path=str(target),
            count=len(documents),
        )

        try:
            df = self._build_dataframe(documents)
            df.to_csv(target, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compatibility
        except Exception as exc:
            dataset_log.error("CSV write failed | path={path} | error={error}", path=str(target), error=str(exc))
            raise DatasetScanError(
                message=f"Failed to write CSV to '{target}': {exc}",
                detail={"path": str(target), "error": str(exc)},
            ) from exc

        dataset_log.info("CSV written successfully | path={path}", path=str(target))
        return target

    def save_documents_xlsx(
        self,
        documents: List[DocumentRecord],
        statistics: DatasetStatistics,
        output_path: Optional[Path] = None,
    ) -> Path:
        """
        Write document records to a professionally formatted Excel file.

        The workbook contains two sheets:
            1. 'Documents' — one row per PDF with full metadata
            2. 'Summary'   — aggregate statistics

        Formatting:
            - Header: navy blue background, white bold font, centered
            - Status column: green (valid), red (invalid/corrupted), yellow (empty/duplicate)
            - All data rows: thin borders, left-aligned text
            - Auto-fit column widths (capped at 50 chars)
            - Frozen header row (scroll while keeping headers visible)

        Args:
            documents:   List of DocumentRecord objects.
            statistics:  DatasetStatistics for the Summary sheet.
            output_path: Target XLSX path. Defaults to metadata/documents.xlsx.

        Returns:
            Resolved absolute path of the written XLSX file.

        Raises:
            DatasetScanError: If writing fails.
        """
        target = (output_path or self._metadata_dir / DOCUMENTS_XLSX_FILENAME).resolve()

        dataset_log.info(
            "Writing XLSX | path={path} | records={count}",
            path=str(target),
            count=len(documents),
        )

        try:
            df = self._build_dataframe(documents)

            # Step 1: Write DataFrame to XLSX using pandas
            with pd.ExcelWriter(str(target), engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Documents")

            # Step 2: Re-open with openpyxl for professional formatting
            workbook = load_workbook(str(target))
            worksheet = workbook["Documents"]

            # Determine Status column index (1-based)
            status_col_idx = self._CSV_COLUMNS.index("Status") + 1

            # Apply formatting
            self._apply_xlsx_header_formatting(worksheet)
            self._apply_xlsx_row_formatting(worksheet, status_col_idx)
            self._apply_xlsx_status_colors(worksheet, status_col_idx)
            self._apply_xlsx_column_widths(worksheet)

            # Freeze the header row
            worksheet.freeze_panes = "A2"

            # Add summary statistics sheet
            self._add_statistics_sheet(workbook, statistics)

            workbook.save(str(target))

        except Exception as exc:
            dataset_log.error("XLSX write failed | path={path} | error={error}", path=str(target), error=str(exc))
            raise DatasetScanError(
                message=f"Failed to write XLSX to '{target}': {exc}",
                detail={"path": str(target), "error": str(exc)},
            ) from exc

        dataset_log.info("XLSX written successfully | path={path}", path=str(target))
        return target

    def save_json_summary(
        self,
        statistics: DatasetStatistics,
        output_path: Optional[Path] = None,
    ) -> Path:
        """
        Write dataset statistics to dataset_summary.json.

        The output format matches the project specification:
        {
            "total_documents": 100,
            "valid_documents": 95,
            ...
            "categories": {"Acts": 35, ...}
        }

        Args:
            statistics:  DatasetStatistics domain model.
            output_path: Target JSON path. Defaults to metadata/dataset_summary.json.

        Returns:
            Resolved absolute path of the written JSON file.

        Raises:
            DatasetScanError: If JSON write fails.
        """
        target = (output_path or self._metadata_dir / "dataset_summary.json").resolve()

        dataset_log.info("Writing JSON summary | path={path}", path=str(target))

        try:
            summary = statistics.to_json_summary()
            with open(target, "w", encoding="utf-8") as json_file:
                json.dump(summary, json_file, indent=2, ensure_ascii=False)
        except Exception as exc:
            dataset_log.error("JSON write failed | path={path} | error={error}", path=str(target), error=str(exc))
            raise DatasetScanError(
                message=f"Failed to write JSON summary to '{target}': {exc}",
                detail={"path": str(target), "error": str(exc)},
            ) from exc

        dataset_log.info("JSON summary written | path={path}", path=str(target))
        return target

    def load_documents_csv(self, csv_path: Optional[Path] = None) -> List[dict]:
        """
        Load previously saved document records from a CSV file.

        Args:
            csv_path: CSV path. Defaults to metadata/documents.csv.

        Returns:
            List of row dicts (pandas to_dict('records')).

        Raises:
            FileNotFoundError: If the CSV does not exist.
            DatasetScanError:  If pandas cannot parse the CSV.
        """
        target = (csv_path or self._metadata_dir / DOCUMENTS_CSV_FILENAME).resolve()

        if not target.exists():
            raise FileNotFoundError(f"Documents CSV not found at '{target}'.")

        try:
            df = pd.read_csv(target, encoding="utf-8-sig")
            return df.to_dict("records")
        except Exception as exc:
            raise DatasetScanError(
                message=f"Failed to read CSV from '{target}': {exc}",
                detail={"path": str(target), "error": str(exc)},
            ) from exc

    def export_all(self, scan_result: ScanResult) -> dict[str, Path]:
        """
        Convenience method: write CSV, XLSX, and JSON in one call.

        Args:
            scan_result: ScanResult from DatasetService.scan_dataset().

        Returns:
            Dict mapping "csv", "xlsx", "json" to their respective output paths.

        Raises:
            DatasetScanError: If any individual export fails.
        """
        csv_path = self.save_documents_csv(scan_result.documents)
        xlsx_path = self.save_documents_xlsx(
            scan_result.documents,
            scan_result.statistics,
        )
        json_path = self.save_json_summary(scan_result.statistics)

        dataset_log.info(
            "All exports complete | csv={csv} | xlsx={xlsx} | json={json}",
            csv=csv_path.name,
            xlsx=xlsx_path.name,
            json=json_path.name,
        )

        return {"csv": csv_path, "xlsx": xlsx_path, "json": json_path}
