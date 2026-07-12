"""
app/golden_set/golden_repository.py
=====================================
Repository (file I/O) for the Golden Set Management Module.

PURPOSE:
    Abstracts all file system operations for the golden set module.
    The service layer calls the repository — it never touches the file
    system directly. This enables easy testing and storage-backend swapping.

    Operations:
        - Save validated records as CSV
        - Save validated records as XLSX (with formatting)
        - Save validation report as JSON
        - Save statistics as JSON
        - Load previously persisted records from CSV

DESIGN:
    - Abstract base class GoldenSetRepository (interface)
    - FileSystemGoldenSetRepository implementation (production)
    - All paths are injected (not hard-coded)
    - No business logic — pure I/O

SOLID:
    Single Responsibility — file persistence only.
    Dependency Inversion  — service depends on abstract base class.
    Open/Closed           — new storage backends without modifying service.
"""

from __future__ import annotations

import csv
import json
from abc import ABC, abstractmethod
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from app.golden_set.golden_logger import golden_log
from app.golden_set.golden_models import (
    GoldenRecord,
    GoldenRecordStatus,
    GoldenSetStatistics,
    ValidationReport,
)


# =============================================================================
# ABSTRACT BASE CLASS
# =============================================================================

class GoldenSetRepository(ABC):
    """
    Abstract base class defining the golden set persistence interface.

    All file I/O implementations must implement these methods.
    The service layer only interacts with this interface.
    """

    @abstractmethod
    def save_validated_csv(self, records: List[GoldenRecord], path: Path) -> None:
        """Save validated GoldenRecord objects as a CSV file."""
        ...

    @abstractmethod
    def save_validated_xlsx(self, records: List[GoldenRecord], path: Path) -> None:
        """Save validated GoldenRecord objects as a formatted XLSX file."""
        ...

    @abstractmethod
    def save_report(self, report: ValidationReport, path: Path) -> None:
        """Persist the ValidationReport as a JSON file."""
        ...

    @abstractmethod
    def save_statistics(self, statistics: GoldenSetStatistics, path: Path) -> None:
        """Persist GoldenSetStatistics as a JSON file."""
        ...

    @abstractmethod
    def load_validated_records(self, path: Path) -> Optional[List[Dict]]:
        """Read back persisted validated records from a CSV file."""
        ...


# =============================================================================
# FILE SYSTEM IMPLEMENTATION
# =============================================================================

class FileSystemGoldenSetRepository(GoldenSetRepository):
    """
    Production file-system implementation of GoldenSetRepository.

    All files are written to the metadata directory:
        metadata/validated_golden_set.csv
        metadata/validated_golden_set.xlsx
        metadata/golden_set_validation_report.json
        metadata/golden_set_statistics.json

    Args:
        metadata_dir: Path to the metadata output directory.
    """

    # CSV columns in export order
    _CSV_COLUMNS: List[str] = [
        "row_number",
        "query",
        "expected_answer",
        "source_document",
        "page_number",
        "category",
        "citation",
        "difficulty",
        "tags",
        "notes",
        "status",
        "query_length",
        "answer_length",
    ]

    def __init__(self, metadata_dir: Path) -> None:
        self._metadata_dir = metadata_dir
        self._metadata_dir.mkdir(parents=True, exist_ok=True)

    def save_validated_csv(self, records: List[GoldenRecord], path: Path) -> None:
        """
        Write validated records to a UTF-8 CSV file.

        Writes ALL records (all statuses) — the status column distinguishes
        valid, invalid, duplicate, and rejected rows.

        Args:
            records: List of GoldenRecord objects.
            path:    Absolute path to write the CSV file.
        """
        golden_log.info(
            "Writing validated CSV | path={path} | records={count}",
            path=str(path),
            count=len(records),
        )
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, "w", encoding="utf-8", newline="") as csvfile:
            writer = csv.DictWriter(
                csvfile,
                fieldnames=self._CSV_COLUMNS,
                extrasaction="ignore",
                quoting=csv.QUOTE_MINIMAL,
            )
            writer.writeheader()
            for record in records:
                if record.status == GoldenRecordStatus.REJECTED:
                    continue  # Skip completely empty/rejected rows
                writer.writerow({
                    "row_number": record.row_number,
                    "query": record.query,
                    "expected_answer": record.expected_answer,
                    "source_document": record.source_document,
                    "page_number": record.page_number,
                    "category": record.category,
                    "citation": record.citation or "",
                    "difficulty": record.difficulty or "",
                    "tags": record.tags or "",
                    "notes": record.notes or "",
                    "status": record.status.value,
                    "query_length": record.query_length,
                    "answer_length": record.answer_length,
                })

        golden_log.info(
            "Validated CSV written | path={path}",
            path=str(path),
        )

    def save_validated_xlsx(self, records: List[GoldenRecord], path: Path) -> None:
        """
        Write validated records to a formatted XLSX file.

        Applies professional formatting: bold header, column widths, and
        status-conditional row coloring (green=valid, red=invalid, orange=duplicate).

        Args:
            records: List of GoldenRecord objects.
            path:    Absolute path to write the XLSX file.
        """
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            golden_log.warning(
                "openpyxl not available; skipping XLSX export to {path}",
                path=str(path),
            )
            return

        golden_log.info(
            "Writing validated XLSX | path={path} | records={count}",
            path=str(path),
            count=len(records),
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validated Golden Set"  # type: ignore[assignment]

        # ── Header row ────────────────────────────────────────────────────────
        header_labels = [
            "Row#", "Query", "Expected Answer", "Source Document",
            "Page", "Category", "Citation", "Difficulty", "Tags", "Notes",
            "Status", "Query Length", "Answer Length",
        ]
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, label in enumerate(header_labels, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)  # type: ignore[union-attr]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Status fill colors ─────────────────────────────────────────────────
        _status_fills: Dict[str, PatternFill] = {
            GoldenRecordStatus.VALID.value: PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
            ),
            GoldenRecordStatus.INVALID.value: PatternFill(
                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
            ),
            GoldenRecordStatus.DUPLICATE.value: PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            ),
        }

        # ── Data rows ─────────────────────────────────────────────────────────
        data_font = Font(size=10)
        row_num = 2
        for record in records:
            if record.status == GoldenRecordStatus.REJECTED:
                continue
            row_data = [
                record.row_number,
                record.query,
                record.expected_answer,
                record.source_document,
                record.page_number,
                record.category,
                record.citation or "",
                record.difficulty or "",
                record.tags or "",
                record.notes or "",
                record.status.value,
                record.query_length,
                record.answer_length,
            ]
            fill = _status_fills.get(record.status.value)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)  # type: ignore[union-attr]
                cell.font = data_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    cell.fill = fill
            row_num += 1

        # ── Column widths ─────────────────────────────────────────────────────
        _col_widths = [8, 50, 60, 25, 8, 20, 30, 12, 30, 30, 12, 14, 15]
        for i, width in enumerate(_col_widths, start=1):
            ws.column_dimensions[  # type: ignore[index]
                openpyxl.utils.get_column_letter(i)  # type: ignore[attr-defined]
            ].width = width

        ws.freeze_panes = "A2"  # type: ignore[assignment]
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        golden_log.info("Validated XLSX written | path={path}", path=str(path))

    def save_report(self, report: ValidationReport, path: Path) -> None:
        """
        Persist the ValidationReport as a formatted JSON file.

        Args:
            report: ValidationReport domain model.
            path:   Absolute path to write the JSON file.
        """
        golden_log.info("Writing validation report | path={path}", path=str(path))
        path.parent.mkdir(parents=True, exist_ok=True)
        report_dict = report.to_summary_dict()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(report_dict, f, indent=2, ensure_ascii=False)
        golden_log.info("Validation report written | path={path}", path=str(path))

    def save_statistics(self, statistics: GoldenSetStatistics, path: Path) -> None:
        """
        Persist GoldenSetStatistics as a formatted JSON file.

        Args:
            statistics: GoldenSetStatistics domain model.
            path:       Absolute path to write the JSON file.
        """
        golden_log.info("Writing statistics | path={path}", path=str(path))
        path.parent.mkdir(parents=True, exist_ok=True)
        stats_dict = statistics.to_json_dict()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(stats_dict, f, indent=2, ensure_ascii=False)
        golden_log.info("Statistics written | path={path}", path=str(path))

    def load_validated_records(self, path: Path) -> Optional[List[Dict]]:
        """
        Read back persisted validated records from a CSV file.

        Args:
            path: Absolute path to the validated_golden_set.csv.

        Returns:
            List of raw row dicts, or None if the file does not exist.
        """
        if not path.exists():
            return None

        rows: List[Dict] = []
        with open(path, "r", encoding="utf-8", newline="") as csvfile:
            reader = csv.DictReader(csvfile)
            rows = list(reader)

        golden_log.info(
            "Loaded {count} validated records from {path}",
            count=len(rows),
            path=str(path),
        )
        return rows
